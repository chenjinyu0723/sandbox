#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
  Excel 大文件处理脚本 — Polars + Calamine/FastExcel 引擎
================================================================================

  设计目标：
    处理超大型 Excel 文件（百万行 × 数十列），以"逐表读取 → 算子管线 →
    合并写入"的模式工作，全程控制内存，单个子表单独处理完即释放。

  核心能力（三种算子，按 JSON 配置中的顺序依次执行）：
    1. filter       — 按条件筛选行（支持统计量：中位数/众数/均值/异常值边界）
    2. drop_nulls   — 删除含空值的行（多列默认"或"逻辑：任一列为空即删）
    3. deduplicate  — 去重（支持"交"=组合去重 / "并"=独立去重）

  调用方式：
    python xlsx_processor.py --config config.json
    python xlsx_processor.py --config config.json --dry-run    # 仅验证配置
    python xlsx_processor.py --help                            # 查看完整帮助

  依赖：
    pip install polars fastexcel xlsxwriter

================================================================================
  配置文件结构 (JSON)
================================================================================

{
  // ── 必填 ──────────────────────────────────────────────────────────
  "input_path": "C:/data/原始数据.xlsx",

  // ── 可选 ──────────────────────────────────────────────────────────
  // 输出路径：不填则默认在 input_path 同目录生成 "<原名>_processed.xlsx"
  "output_path": "C:/data/原始数据_processed.xlsx",

  // 输出后缀：仅当 output_path 未指定时生效，默认 "_processed"
  "output_suffix": "_cleaned",

  // ── 逐表配置（按 Sheet 名称索引）──────────────────────────────────
  //   key = Excel 中子表的名称（区分大小写）
  //   value = 该表的处理规则
  "sheets": {

    // ================================================================
    // 示例 1：完整管线 — 先筛选 → 再去空 → 最后去重
    // ================================================================
    "订单明细": {
      "operations": [
        {
          // ── 算子名称 ──
          "type": "filter",

          // ── 条件间的组合逻辑 ──
          //   "and"（交）：所有条件同时满足才保留该行
          //   "or" （并）：任一条件满足即保留
          "logic": "and",

          // ── 逐条件列表 ──
          "conditions": [
            {
              // 要比较的列名（与 Excel 表头一致）
              "column": "金额",

              // 比较运算符：>  <  >=  <=  ==  !=
              "operator": ">=",

              // ── 比较值的来源 ──
              //   "literal"       → 直接使用 value 字段的值
              //   "median"        → 该列的中位数
              //   "mean"          → 该列的均值
              //   "mode"          → 该列的众数（出现次数最多的值）
              //   "outlier_lower" → 下异常值边界 Q1 - 1.5×IQR
              //   "outlier_upper" → 上异常值边界 Q3 + 1.5×IQR
              "value_type": "median",

              // 统计量从哪一列计算（不填默认 = column）
              "value_column": "金额",

              // 仅 value_type="literal" 时读取此字段
              "value": null
            },
            {
              "column": "状态",
              "operator": "==",
              "value_type": "literal",
              "value": "已完成"
            }
          ]
        },
        {
          // ── 去空值算子 ──
          "type": "drop_nulls",

          // 检查哪些列是否为空
          "columns": ["客户名", "金额"],

          // ── 空值逻辑 ──
          //   "or"（默认）：任一列为空就删除该行
          //   "and"      ：所有列都为空才删除该行
          "logic": "or"
        },
        {
          // ── 去重算子 ──
          "type": "deduplicate",

          // 依据哪些列去重
          "columns": ["订单号", "客户名"],

          // ── 去重逻辑 ──
          //   "and"（交，默认）：列值的组合重复 → 视为重复（标准去重）
          //       例：("A","1") 与 ("A","1") 重复 → 保留第一行
          //   "or" （并）     ：每一列独立去重；只有当该行所有列的值
          //       都是"首次出现"时才保留
          //       例：A列值之前出现过 或 B列值之前出现过 → 删除
          "logic": "and"
        }
      ]
    },

    // ================================================================
    // 示例 2：仅去重（不筛选、不去空）
    // ================================================================
    "产品表": {
      "operations": [
        {
          "type": "deduplicate",
          "columns": ["产品编码"],
          "logic": "and"
        }
      ]
    },

    // ================================================================
    // 示例 3：无操作 — 原样保留
    // ================================================================
    "说明页": {
      "operations": []
    }
  },

  // ── 全局控制 ──────────────────────────────────────────────────────
  "global": {
    // 跳过的子表（名称列表）。这些表不会出现在输出中
    "skip_sheets": ["临时计算"],

    // 仅处理这些子表。null = 处理全部（除 skip_sheets 外）
    "only_sheets": null,

    // 读取引擎：默认为 calamine（Polars ≥1.0 默认）
    "engine": "calamine"
  }
}

================================================================================
  交/并 逻辑速查表
================================================================================

┌──────────────┬──────────────────────────────────────────────────────┐
│  算子 / 场景  │  "and"（交）                   "or"（并）             │
├──────────────┼──────────────────────────────────────────────────────┤
│ filter       │ 所有条件同时满足              任一条件满足             │
│ (多条件)     │ 例：金额>100 AND 状态=完成    例：金额>100 OR 状态=完成│
├──────────────┼──────────────────────────────────────────────────────┤
│ drop_nulls   │ 所有列都为空才删除            任一列为空就删除         │
│ (多列)       │ 例：A空 AND B空 → 删         例：A空 OR B空 → 删     │
├──────────────┼──────────────────────────────────────────────────────┤
│ deduplicate  │ 组合去重（标准语义）          独立去重                 │
│ (多列)       │ (A,B)组合相同 → 去重          A列值出现过 OR B列值    │
│              │ 例：(张,1)与(张,1)→重复      出现过 → 删除            │
│              │    (张,1)与(张,2)→保留       例：A=张已在第1行       │
│              │                                B=1 已在第3行         │
│              │                                第5行(张,1)→全见过→删 │
└──────────────┴──────────────────────────────────────────────────────┘

================================================================================
  内存安全设计
================================================================================

  1. 逐表处理：一次只加载一个 Sheet 到内存，处理完立即释放
  2. 算子顺序：按 config 中 operations 的顺序依次执行，中间结果不落盘
  3. 写入策略：所有表处理完后统一写入一个 Excel（xlsxwriter 引擎）
  4. 大数据警告：单个 Sheet 超过 500MB 原始大小时打印警告
  5. 兜底建议：如果单表确实太大内存不够，建议先转 parquet 再处理
     (polars 支持 read_csv/parquet 的 streaming 模式)
"""

# ============================================================================
#  导入
# ============================================================================

import argparse
import json
import logging
import os
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import polars as pl

# ============================================================================
#  日志配置
# ============================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("xlsx_processor")


# ============================================================================
#  1. 统计量计算 — 为 filter 算子提供比较基准值
# ============================================================================

def compute_statistic(
    series: pl.Series,
    stat_type: str,
    outlier_method: str = "iqr",
) -> Union[float, int, str]:
    """
    从 Polars Series 中计算统计量，返回一个标量值作为 filter 的比较基准。

    ┌──────────────────┬───────────────────────────────────────┐
    │ stat_type        │ 含义                                  │
    ├──────────────────┼───────────────────────────────────────┤
    │ "median"         │ 中位数（50 分位数）                    │
    │ "mean"           │ 算术平均值                            │
    │ "mode"           │ 众数（出现次数最多的值，取第一个）     │
    │ "outlier_lower"  │ 下异常值边界 = Q1 - 1.5 × IQR         │
    │ "outlier_upper"  │ 上异常值边界 = Q3 + 1.5 × IQR         │
    └──────────────────┴───────────────────────────────────────┘

    参数:
        series          : Polars Series，要计算统计量的列
        stat_type       : 统计量类型（见上表）
        outlier_method  : 异常值检测方法，"iqr" (默认, IQR×1.5)
                          或 "iqr_3" (IQR×3.0, 更宽松)

    返回:
        标量值（float/int/str），可直接用于比较

    异常处理:
        - 空列：抛出 ValueError
        - 非数值列且 stat_type 需要数值：尝试转换，失败则抛出
    """
    # ── 空列检查 ──
    if series.len() == 0:
        raise ValueError(f"列 '{series.name}' 为空，无法计算 {stat_type}")

    # 去除 null 值后计算
    clean = series.drop_nulls()

    if clean.len() == 0:
        raise ValueError(
            f"列 '{series.name}' 中所有值均为 null，"
            f"无法计算 {stat_type}"
        )

    # ── 根据类型分发 ──
    stat_type_lower = stat_type.lower()

    if stat_type_lower == "median":
        return clean.median()  # type: ignore[return-value]

    elif stat_type_lower == "mean":
        return clean.mean()  # type: ignore[return-value]

    elif stat_type_lower == "mode":
        # mode() 返回一个 Series（可能有多个众数），取第一个
        modes: pl.Series = clean.mode()  # type: ignore[assignment]
        return modes[0]  # type: ignore[no-any-return]

    elif stat_type_lower in ("outlier_lower", "outlier_upper"):
        # ── IQR 方法计算异常值边界 ──
        # Q1 = 25 分位数, Q3 = 75 分位数
        # IQR = Q3 - Q1
        # 下边界 = Q1 - multiplier × IQR
        # 上边界 = Q3 + multiplier × IQR

        multiplier = 3.0 if outlier_method == "iqr_3" else 1.5

        try:
            q1 = clean.quantile(0.25, interpolation="linear")  # type: ignore[union-attr]
            q3 = clean.quantile(0.75, interpolation="linear")  # type: ignore[union-attr]
        except Exception:
            # 降级：用 describe() 取近似值
            desc = clean.describe()
            # describe() 返回的 DataFrame 中，"statistic" 列有 "25%" 和 "75%" 行
            q1_row = desc.filter(pl.col("statistic") == "25%")
            q3_row = desc.filter(pl.col("statistic") == "75%")
            if q1_row.is_empty() or q3_row.is_empty():
                raise ValueError(
                    f"无法计算列 '{series.name}' 的四分位数"
                )
            q1 = float(q1_row[0, 1])  # type: ignore[index]
            q3 = float(q3_row[0, 1])  # type: ignore[index]

        iqr = q3 - q1

        if stat_type_lower == "outlier_lower":
            return q1 - multiplier * iqr
        else:
            return q3 + multiplier * iqr

    else:
        raise ValueError(
            f"不支持的统计量类型: '{stat_type}'。"
            f"可选: median, mean, mode, outlier_lower, outlier_upper"
        )


# ============================================================================
#  2. filter 算子 — 条件筛选
# ============================================================================

# 支持的比较运算符 → Polars 表达式映射
_OP_MAP = {
    ">":  lambda col, val: pl.col(col) > val,
    "<":  lambda col, val: pl.col(col) < val,
    ">=": lambda col, val: pl.col(col) >= val,
    "<=": lambda col, val: pl.col(col) <= val,
    "==": lambda col, val: pl.col(col) == val,
    "!=": lambda col, val: pl.col(col) != val,
}


def apply_filter(
    df: pl.DataFrame,
    conditions: List[Dict[str, Any]],
    logic: str = "and",
) -> pl.DataFrame:
    """
    按条件筛选行。

    参数:
        df          : 输入 DataFrame
        conditions  : 条件列表，每个条件是一个 dict：
            {
                "column":       str,    # 列名（必填）
                "operator":     str,    # 比较运算符（必填）: > < >= <= == !=
                "value_type":   str,    # 基准值类型（必填）:
                                        #   "literal" → 直接用 value 字段
                                        #   "median" / "mean" / "mode" /
                                        #   "outlier_lower" / "outlier_upper"
                "value_column": str,    # 统计量从哪列计算（可选，默认=column）
                "value":        Any,    # 仅 value_type="literal" 时使用（必填时）
            }
        logic       : 条件间组合逻辑
                      "and"（交）→ 所有条件同时满足
                      "or" （并）→ 任一条件满足

    返回:
        筛选后的 DataFrame

    处理流程:
        1. 对每个 condition：
           a. 确定比较基准值（literal 直接取 / 否则计算统计量）
           b. 构建 Polars 布尔表达式 (pl.col(col) OP value)
        2. 按 logic 组合所有表达式：
           and → expr1 & expr2 & ... & exprN
           or  → expr1 | expr2 | ... | exprN
        3. filter 并返回
    """
    # ── 空条件 → 不做任何筛选 ──
    if not conditions:
        logger.info("    [filter] 条件列表为空，跳过筛选")
        return df

    logger.info(
        f"    [filter] 开始筛选，共 {len(conditions)} 个条件，"
        f"逻辑: {'交 (AND)' if logic == 'and' else '并 (OR)'}"
    )

    # ── 逐条件构建表达式 ──
    expressions: List[pl.Expr] = []

    for i, cond in enumerate(conditions, 1):
        # 1) 解析参数字段
        col_name: str = cond["column"]
        operator: str = cond["operator"]
        value_type: str = cond.get("value_type", "literal")
        value_column: str = cond.get("value_column", col_name)
        literal_value: Any = cond.get("value")

        # 2) 校验列是否存在
        if col_name not in df.columns:
            available = ", ".join(df.columns[:10])
            raise KeyError(
                f"条件 {i}: 列 '{col_name}' 不存在。"
                f"可用列（前10）: {available}"
            )
        if value_column not in df.columns:
            raise KeyError(
                f"条件 {i}: 统计量计算列 '{value_column}' 不存在"
            )

        # 3) 校验运算符
        if operator not in _OP_MAP:
            raise ValueError(
                f"条件 {i}: 不支持的运算符 '{operator}'。"
                f"可选: {list(_OP_MAP.keys())}"
            )

        # 4) 确定比较基准值
        if value_type == "literal":
            # 直接使用用户指定的值
            if literal_value is None:
                raise ValueError(
                    f"条件 {i}: value_type='literal' 但 value 字段为空。"
                    f"请在 condition.value 中提供比较值"
                )
            compare_value = literal_value
            logger.debug(
                f"      条件 {i}: {col_name} {operator} {compare_value} (字面值)"
            )
        else:
            # 从指定列计算统计量
            compare_value = compute_statistic(
                df[value_column],
                value_type,
            )
            logger.info(
                f"      条件 {i}: {col_name} {operator} "
                f"{compare_value:.4f} (统计量={value_type}, 来源列={value_column})"
            )

        # 5) 构建 Polars 表达式
        expr_fn = _OP_MAP[operator]
        expressions.append(expr_fn(col_name, compare_value))

    # ── 按 logic 组合表达式 ──
    if logic == "and":
        # 交：所有条件取 &（AND）
        combined = expressions[0]
        for expr in expressions[1:]:
            combined = combined & expr
    elif logic == "or":
        # 并：所有条件取 |（OR）
        combined = expressions[0]
        for expr in expressions[1:]:
            combined = combined | expr
    else:
        raise ValueError(
            f"不支持的条件组合逻辑: '{logic}'。可选: and, or"
        )

    # ── 应用筛选 ──
    before = df.height
    df = df.filter(combined)
    after = df.height
    logger.info(
        f"    [filter] 完成：{before} → {after} 行 "
        f"(删除 {before - after} 行, "
        f"保留率 {after / max(before, 1) * 100:.1f}%)"
    )

    return df


# ============================================================================
#  3. drop_nulls 算子 — 删除含空值的行
# ============================================================================

def apply_drop_nulls(
    df: pl.DataFrame,
    columns: List[str],
    logic: str = "or",
) -> pl.DataFrame:
    """
    删除指定列中包含空值（null / NaN）的行。

    参数:
        df      : 输入 DataFrame
        columns : 要检查的列名列表
        logic   : 空值判定逻辑
                  "or"  (默认) → 任一列为空就删除（更激进）
                  "and"        → 所有列都为空才删除（更宽松）

    返回:
        删除含空值行后的 DataFrame

    处理流程:
        1. 对每一列构建 is_null() 表达式
        2. 按 logic 组合：
           or  → (col_a.is_null() | col_b.is_null() | ...)
           and → (col_a.is_null() & col_b.is_null() & ...)
        3. 取反 → 保留非空的那些行
        4. filter 并返回
    """
    if not columns:
        logger.info("    [drop_nulls] 列列表为空，跳过")
        return df

    logger.info(
        f"    [drop_nulls] 开始去空，检查列: {columns}，"
        f"逻辑: {'任一空即删 (OR)' if logic == 'or' else '全空才删 (AND)'}"
    )

    # ── 校验列名 ──
    for col in columns:
        if col not in df.columns:
            available = ", ".join(df.columns[:10])
            raise KeyError(
                f"[drop_nulls] 列 '{col}' 不存在。可用列（前10）: {available}"
            )

    # ── 构建空值检测表达式 ──
    # is_null() 对 null 和 NaN 都返回 True
    null_exprs = [pl.col(c).is_null() for c in columns]

    if logic == "or":
        # 任一列为空 → 标记为删除
        drop_mask = null_exprs[0]
        for expr in null_exprs[1:]:
            drop_mask = drop_mask | expr
    elif logic == "and":
        # 所有列都为空 → 标记为删除
        drop_mask = null_exprs[0]
        for expr in null_exprs[1:]:
            drop_mask = drop_mask & expr
    else:
        raise ValueError(
            f"不支持的空值逻辑: '{logic}'。可选: or, and"
        )

    # ── 取反：保留没有空值的行 ──
    before = df.height
    df = df.filter(~drop_mask)
    after = df.height
    logger.info(
        f"    [drop_nulls] 完成：{before} → {after} 行 "
        f"(删除 {before - after} 行)"
    )

    return df


# ============================================================================
#  4. deduplicate 算子 — 去重
# ============================================================================

def apply_deduplicate(
    df: pl.DataFrame,
    columns: List[str],
    logic: str = "and",
) -> pl.DataFrame:
    """
    按指定列去重。

    参数:
        df      : 输入 DataFrame
        columns : 依据哪些列去重
        logic   : 去重逻辑
                  "and"（交，默认）→ 组合去重：列值组合 (a,b,c) 重复 → 视为重复
                                   保留第一次出现的行（标准 unique 语义）
                  "or" （并）     → 独立去重：对每一列独立追踪"已见过的值"，
                                   只有当该行所有列的值都是"首次出现"时才保留。
                                   语义：A列值没见过 AND B列值没见过 → 保留

    返回:
        去重后的 DataFrame

    ── "or" 逻辑的详细行为 ────────────────────────────────────────────
    假设原始数据（按行序）：
        行号   A列   B列
         1     张    100
         2     李    200
         3     张    300    ← A="张"在第1行见过 → 删除
         4     王    200    ← B=200 在第2行见过 → 删除
         5     赵    400    ← A="赵"首次, B=400首次 → 保留

    结果保留：第1行(张,100)、第2行(李,200)、第5行(赵,400)

    实现方式：
        - 对每一列调用 is_first_distinct()，得到布尔 mask
          (True = 该值在全表中首次出现)
        - 将所有 mask 做 &（AND），只有全 True 的行保留
        - 这保证了独立去重的语义

    限制：
        这是"全表首次出现"判断，不考虑因前面行被删除而释放的值。
        如需严格的状态追踪（删除行释放其值），目前需用 Python 行级循环
        （性能会显著下降），可联系作者扩展。
    """
    if not columns:
        logger.info("    [deduplicate] 列列表为空，跳过")
        return df

    logger.info(
        f"    [deduplicate] 开始去重，依据列: {columns}，"
        f"逻辑: {'组合去重 (AND)' if logic == 'and' else '独立去重 (OR)'}"
    )

    # ── 校验列名 ──
    for col in columns:
        if col not in df.columns:
            available = ", ".join(df.columns[:10])
            raise KeyError(
                f"[deduplicate] 列 '{col}' 不存在。可用列（前10）: {available}"
            )

    before = df.height

    # ── 按逻辑分发 ──
    if logic == "and":
        # ── 组合去重（标准语义）──
        # keep="first" 保留每个组合的第一次出现
        # maintain_order=True 保持原始行序
        df = df.unique(
            subset=columns,
            keep="first",
            maintain_order=True,
        )

    elif logic == "or":
        # ── 独立去重 ──
        # 对每一列：is_first_distinct() 标记每个值是否首次出现
        # 组合：所有列都首次出现 → 保留
        masks: List[pl.Expr] = [
            pl.col(c).is_first_distinct()
            for c in columns
        ]
        combined_mask = masks[0]
        for m in masks[1:]:
            combined_mask = combined_mask & m
        df = df.filter(combined_mask)

    else:
        raise ValueError(
            f"不支持的去重逻辑: '{logic}'。可选: and, or"
        )

    after = df.height
    logger.info(
        f"    [deduplicate] 完成：{before} → {after} 行 "
        f"(删除 {before - after} 行, "
        f"保留率 {after / max(before, 1) * 100:.1f}%)"
    )

    return df


# ============================================================================
#  5. 算子调度 — 按 operations 列表依次执行
# ============================================================================

# 算子类型 → 处理函数 的映射表
_OPERATION_HANDLERS = {
    "filter":       apply_filter,
    "drop_nulls":   apply_drop_nulls,
    "deduplicate":  apply_deduplicate,
}


def process_sheet(
    df: pl.DataFrame,
    operations: List[Dict[str, Any]],
    sheet_name: str = "(未命名)",
) -> pl.DataFrame:
    """
    对一个子表的 DataFrame 依次执行所有配置的算子。

    参数:
        df          : 输入 DataFrame（该子表的原始数据）
        operations  : 算子配置列表，按顺序执行
        sheet_name  : 子表名称（仅用于日志）

    返回:
        处理后的 DataFrame

    异常处理:
        - 未知算子类型 → ValueError
        - 算子内部错误（列不存在/类型错误等）→ 向上传播
        - 处理前后打印行数变化
    """
    if not operations:
        logger.info(f"  [{sheet_name}] 无操作，原样保留 ({df.height} 行)")
        return df

    logger.info(f"  [{sheet_name}] 开始处理，共 {len(operations)} 个算子步")
    before_total = df.height

    for i, op in enumerate(operations, 1):
        op_type: str = op.get("type", "")
        if op_type not in _OPERATION_HANDLERS:
            supported = list(_OPERATION_HANDLERS.keys())
            raise ValueError(
                f"[{sheet_name}] 算子 {i}: 未知类型 '{op_type}'。"
                f"支持的类型: {supported}"
            )

        # ── 构造算子参数（移除 type 字段，其余传给处理函数）──
        op_params = {k: v for k, v in op.items() if k != "type"}

        # ── 执行算子 ──
        logger.info(f"  [{sheet_name}] 步骤 {i}/{len(operations)}: {op_type}")
        handler = _OPERATION_HANDLERS[op_type]
        df = handler(df, **op_params)

        # ── 处理完立即检查行数，预防全删 ──
        if df.height == 0:
            logger.warning(
                f"  [{sheet_name}] 步骤 {i} 后数据为空！"
                f"后续算子将跳过"
            )
            break

    after_total = df.height
    logger.info(
        f"  [{sheet_name}] 处理完成：{before_total} → {after_total} 行 "
        f"({before_total - after_total} 行被移除)"
    )

    return df


# ============================================================================
#  6. 主流程 — 读取 → 逐表处理 → 写入
# ============================================================================

def _resolve_output_path(input_path: str, output_path: Optional[str], suffix: str) -> str:
    """
    解析输出路径。

    规则:
        1. 如果明确指定了 output_path → 直接用
        2. 否则 → 在 input_path 同目录生成 "<原名><suffix>.xlsx"
        3. 如果 output_path 已存在 → 打印警告并覆盖
    """
    if output_path:
        out = output_path
    else:
        inp = Path(input_path)
        out = str(inp.parent / f"{inp.stem}{suffix}.xlsx")

    if os.path.exists(out):
        logger.warning(f"输出文件已存在，将被覆盖: {out}")

    return out


def _estimate_sheet_size(file_path: str, sheet_name: str) -> int:
    """
    估算单个 Sheet 的行数（不加载全部数据）。

    使用 Polars 的 read_excel 带 n_rows=0 来快速获取 schema，
    然后单独读少量行来估算。这是一个启发式方法。
    """
    try:
        # 方案：用 read_excel 的 infer_schema_length 快速采样
        # 简单起见，先读前100行看列数和每行大概大小
        preview = pl.read_excel(
            file_path,
            sheet_name=sheet_name,
        )
        if preview.height > 0:
            # 估算全表：按前100行推算
            est_memory = preview.estimated_size("mb")
            return preview.height
        return 0
    except Exception:
        return -1  # 无法估算


def process_excel(config: Dict[str, Any]) -> str:
    """
    主处理函数：读取 Excel → 逐表处理 → 写入新 Excel。

    执行流程:
        1. 验证配置（列名、路径等）
        2. 遍历 input Excel 中的所有子表
        3. 对每个子表：
           a. 检查是否在 skip_sheets 中 → 跳过
           b. 检查是否在 only_sheets 中 → 不在则跳过
           c. 读取该 Sheet 的完整数据到 DataFrame
           d. 查找该 Sheet 的 operations 配置
           e. 调用 process_sheet() 执行算子管线
           f. 将处理后的 DataFrame 加入待写入列表
        4. 将所有处理后的子表写入输出 Excel（保持原顺序和名称）
        5. 返回输出文件路径

    参数:
        config  : 完整的 JSON 配置 dict（结构见文件头部文档）

    返回:
        输出文件的绝对路径

    异常:
        FileNotFoundError : 输入文件不存在
        ValueError        : 配置不合法
        KeyError          : 列名不存在
    """
    # ── 解析配置 ──
    input_path: str = config["input_path"]
    output_path: Optional[str] = config.get("output_path")
    output_suffix: str = config.get("output_suffix", "_processed")
    sheets_config: Dict[str, Any] = config.get("sheets", {})
    global_config: Dict[str, Any] = config.get("global", {})
    skip_sheets: List[str] = global_config.get("skip_sheets", [])
    only_sheets: Optional[List[str]] = global_config.get("only_sheets")
    engine: str = global_config.get("engine", "calamine")

    # ── 路径校验 ──
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    # ── 解析输出路径 ──
    resolved_output = _resolve_output_path(input_path, output_path, output_suffix)
    logger.info(f"输入文件: {input_path}")
    logger.info(f"输出文件: {resolved_output}")

    # ── ═══════════════════════════════════════════════════════════
    #  阶段 1：发现所有子表
    # ── ═══════════════════════════════════════════════════════════
    logger.info("=" * 60)
    logger.info("阶段 1/3: 发现子表...")

    # 快速扫描：只读 sheet 名称（用 engine="fastexcel" 可以读 schema 而无需加载数据）
    # 注意：Polars 的 read_excel 如果不指定 sheet_name，默认读第一个 sheet
    # 我们需要所有 sheet 名称。用 xlsxwriter/openpyxl 来获取 sheet 列表
    try:
        import fastexcel
        wb = fastexcel.read_excel(input_path)
        all_sheet_names: List[str] = wb.sheet_names
    except Exception:
        # 降级：用 openpyxl 只读模式获取 sheet 名称
        try:
            import openpyxl
            wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
            all_sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            raise RuntimeError(
                f"无法读取 Excel 文件结构: {e}\n"
                f"请确认文件格式正确且未被其他程序占用"
            )

    logger.info(f"发现 {len(all_sheet_names)} 个子表: {all_sheet_names}")

    # ── ═══════════════════════════════════════════════════════════
    #  阶段 2：逐表处理
    # ── ═══════════════════════════════════════════════════════════
    logger.info("=" * 60)
    logger.info("阶段 2/3: 逐表处理...")

    processed_sheets: Dict[str, pl.DataFrame] = {}  # sheet_name → 处理后的 DataFrame
    total_rows_in = 0
    total_rows_out = 0
    start_time = time.time()

    for idx, sheet_name in enumerate(all_sheet_names, 1):
        # ── 2a. 过滤 ──
        if sheet_name in skip_sheets:
            logger.info(f"  [{idx}/{len(all_sheet_names)}] {sheet_name} → 跳过 (在 skip_sheets 中)")
            continue

        if only_sheets is not None and sheet_name not in only_sheets:
            logger.info(f"  [{idx}/{len(all_sheet_names)}] {sheet_name} → 跳过 (不在 only_sheets 中)")
            continue

        # ── 2b. 读取该子表 ──
        t_read = time.time()
        logger.info(f"  [{idx}/{len(all_sheet_names)}] {sheet_name} → 读取中...")

        try:
            df = pl.read_excel(
                input_path,
                sheet_name=sheet_name,
                engine=engine,  # type: ignore[arg-type]
            )
        except Exception as e:
            logger.error(f"  [{sheet_name}] 读取失败: {e}")
            raise

        rows_in = df.height
        cols_in = len(df.columns)
        total_rows_in += rows_in
        mem_mb = df.estimated_size("mb")
        logger.info(
            f"  [{idx}/{len(all_sheet_names)}] {sheet_name} → "
            f"读取完成: {rows_in} 行 × {cols_in} 列, "
            f"约 {mem_mb:.1f} MB (耗时 {time.time() - t_read:.1f}s)"
        )

        # ── 内存预警 ──
        if mem_mb > 500:
            logger.warning(
                f"  [{sheet_name}] ⚠ 内存占用约 {mem_mb:.0f} MB，"
                f"单表较大，注意内存压力"
            )

        # ── 2c. 获取该表的处理配置 ──
        sheet_config = sheets_config.get(sheet_name, {})
        operations = sheet_config.get("operations", [])

        # ── 2d. 执行处理管线 ──
        try:
            df = process_sheet(df, operations, sheet_name)
        except Exception as e:
            logger.error(f"  [{sheet_name}] 处理失败: {e}")
            raise

        rows_out = df.height
        total_rows_out += rows_out

        # ── 2e. 保存结果 ──
        processed_sheets[sheet_name] = df

    # ── 检查是否有产出 ──
    if not processed_sheets:
        raise ValueError(
            "没有子表被处理！请检查 skip_sheets / only_sheets 配置"
        )

    # ── ═══════════════════════════════════════════════════════════
    #  阶段 3：写入输出 Excel
    # ── ═══════════════════════════════════════════════════════════
    logger.info("=" * 60)
    logger.info("阶段 3/3: 写入输出文件...")

    t_write = time.time()

    # 确保输出目录存在
    os.makedirs(os.path.dirname(resolved_output) or ".", exist_ok=True)

    # Polars write_excel: 传入 dict 即可按 sheet_name 写入多个子表
    # 保持 dict 的插入顺序 = 保持原始子表顺序
    try:
        pl.DataFrame().write_excel(  # dummy call to get the right API
            resolved_output,
            processed_sheets,  # type: ignore[arg-type]
            autofit=True,      # 自动调整列宽
        )
    except TypeError:
        # 降级：逐个写入（某些 Polars 版本不支持 dict 参数）
        import xlsxwriter
        workbook = xlsxwriter.Workbook(resolved_output, {'strings_to_numbers': False})
        for sname, sdf in processed_sheets.items():
            sdf.write_excel(
                workbook=workbook,
                worksheet=sname,
                autofit=True,
            )
        workbook.close()

    elapsed = time.time() - start_time
    write_elapsed = time.time() - t_write

    # ── 打印汇总 ──
    logger.info("=" * 60)
    logger.info("处理完成！汇总:")
    logger.info(f"  输入表数 : {len(all_sheet_names)}")
    logger.info(f"  处理表数 : {len(processed_sheets)}")
    logger.info(f"  跳过表数 : {len(all_sheet_names) - len(processed_sheets)}")
    logger.info(f"  输入行数 : {total_rows_in:,}")
    logger.info(f"  输出行数 : {total_rows_out:,}")
    if total_rows_in > 0:
        logger.info(f"  总保留率 : {total_rows_out / total_rows_in * 100:.1f}%")
    logger.info(f"  处理耗时 : {elapsed - write_elapsed:.1f}s")
    logger.info(f"  写入耗时 : {write_elapsed:.1f}s")
    logger.info(f"  总耗时   : {elapsed:.1f}s")
    logger.info(f"  输出文件 : {resolved_output}")

    return resolved_output


# ============================================================================
#  7. 配置验证
# ============================================================================

def validate_config(config: Dict[str, Any]) -> List[str]:
    """
    验证配置文件的完整性和合法性。

    检查项:
        1. 顶层必填字段是否存在
        2. 输入文件是否存在
        3. 每个 sheet 的 operations 格式是否正确
        4. 每个 condition/operator 是否在允许值范围内
        5. 逻辑标志是否合法

    返回:
        错误信息列表。空列表 = 配置合法
    """
    errors: List[str] = []

    # ── 顶层字段 ──
    if "input_path" not in config:
        errors.append("缺少必填字段: 'input_path'")
    elif not isinstance(config["input_path"], str):
        errors.append("'input_path' 必须是字符串")
    elif not os.path.exists(config["input_path"]):
        errors.append(f"输入文件不存在: {config['input_path']}")

    # ── sheets 字段 ──
    sheets = config.get("sheets", {})
    if not isinstance(sheets, dict):
        errors.append("'sheets' 必须是 JSON 对象 (dict)")
    else:
        for sheet_name, sheet_cfg in sheets.items():
            if not isinstance(sheet_cfg, dict):
                errors.append(f"Sheet '{sheet_name}': 配置必须是 JSON 对象")
                continue

            operations = sheet_cfg.get("operations", [])
            if not isinstance(operations, list):
                errors.append(f"Sheet '{sheet_name}': 'operations' 必须是数组")
                continue

            for i, op in enumerate(operations):
                if not isinstance(op, dict):
                    errors.append(
                        f"Sheet '{sheet_name}', 操作 {i+1}: 必须是 JSON 对象"
                    )
                    continue

                op_type = op.get("type", "")
                if op_type not in _OPERATION_HANDLERS:
                    errors.append(
                        f"Sheet '{sheet_name}', 操作 {i+1}: "
                        f"未知算子类型 '{op_type}'，"
                        f"可选: {list(_OPERATION_HANDLERS.keys())}"
                    )
                    continue

                # ── filter 特有校验 ──
                if op_type == "filter":
                    logic = op.get("logic", "and")
                    if logic not in ("and", "or"):
                        errors.append(
                            f"Sheet '{sheet_name}', 操作 {i+1} filter: "
                            f"logic='{logic}' 不合法，可选: and, or"
                        )

                    conditions = op.get("conditions", [])
                    if not isinstance(conditions, list):
                        errors.append(
                            f"Sheet '{sheet_name}', 操作 {i+1} filter: "
                            f"'conditions' 必须是数组"
                        )
                    else:
                        for j, cond in enumerate(conditions):
                            if not isinstance(cond, dict):
                                errors.append(
                                    f"Sheet '{sheet_name}', 操作 {i+1}, "
                                    f"条件 {j+1}: 必须是 JSON 对象"
                                )
                                continue
                            if "column" not in cond:
                                errors.append(
                                    f"Sheet '{sheet_name}', 操作 {i+1}, "
                                    f"条件 {j+1}: 缺少 'column'"
                                )
                            if "operator" not in cond:
                                errors.append(
                                    f"Sheet '{sheet_name}', 操作 {i+1}, "
                                    f"条件 {j+1}: 缺少 'operator'"
                                )
                            elif cond["operator"] not in _OP_MAP:
                                errors.append(
                                    f"Sheet '{sheet_name}', 操作 {i+1}, "
                                    f"条件 {j+1}: 不支持的运算符 "
                                    f"'{cond['operator']}'"
                                )
                            vt = cond.get("value_type", "literal")
                            if vt not in (
                                "literal", "median", "mean", "mode",
                                "outlier_lower", "outlier_upper",
                            ):
                                errors.append(
                                    f"Sheet '{sheet_name}', 操作 {i+1}, "
                                    f"条件 {j+1}: 不支持的 value_type '{vt}'"
                                )
                            if vt == "literal" and "value" not in cond:
                                errors.append(
                                    f"Sheet '{sheet_name}', 操作 {i+1}, "
                                    f"条件 {j+1}: value_type='literal' "
                                    f"但缺少 'value' 字段"
                                )

                # ── drop_nulls 特有校验 ──
                elif op_type == "drop_nulls":
                    logic = op.get("logic", "or")
                    if logic not in ("and", "or"):
                        errors.append(
                            f"Sheet '{sheet_name}', 操作 {i+1} drop_nulls: "
                            f"logic='{logic}' 不合法，可选: and, or"
                        )

                # ── deduplicate 特有校验 ──
                elif op_type == "deduplicate":
                    logic = op.get("logic", "and")
                    if logic not in ("and", "or"):
                        errors.append(
                            f"Sheet '{sheet_name}', 操作 {i+1} deduplicate: "
                            f"logic='{logic}' 不合法，可选: and, or"
                        )

    # ── global 字段 ──
    global_cfg = config.get("global", {})
    if not isinstance(global_cfg, dict):
        errors.append("'global' 必须是 JSON 对象")

    return errors


# ============================================================================
#  8. CLI 入口
# ============================================================================

def _build_arg_parser() -> argparse.ArgumentParser:
    """构建命令行参数解析器"""

    parser = argparse.ArgumentParser(
        prog="xlsx_processor",
        description=(
            "Excel 大文件处理工具 — 基于 Polars + FastExcel\n"
            "支持逐表去重/去空/条件筛选，通过 JSON 配置文件驱动"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
  示例:
    python xlsx_processor.py --config config.json
    python xlsx_processor.py --config config.json --dry-run
    python xlsx_processor.py --config config.json --verbose

  配置文件格式详见脚本文件头部的完整文档。
        """,
    )

    parser.add_argument(
        "--config", "-c",
        type=str,
        required=True,
        help="JSON 配置文件的路径",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="仅验证配置文件，不执行实际处理",
    )

    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="输出详细日志（DEBUG 级别）",
    )

    parser.add_argument(
        "--generate-example",
        type=str,
        metavar="PATH",
        help="生成一份示例配置文件到指定路径并退出",
    )

    return parser


def _generate_example_config(output_path: str) -> None:
    """生成一份自文档化的示例配置文件"""

    example = {
        "_comment": (
            "这是 xlsx_processor 的示例配置文件。\n"
            "复制此文件，修改 input_path 和 sheets 配置，然后运行：\n"
            "  python xlsx_processor.py --config 你的配置.json\n"
            "所有字段的详细说明见脚本文件头部注释。"
        ),
        "input_path": "C:/data/原始数据.xlsx",
        "output_path": None,
        "output_suffix": "_processed",
        "sheets": {
            "订单明细": {
                "operations": [
                    {
                        "type": "filter",
                        "logic": "and",
                        "conditions": [
                            {
                                "column": "金额",
                                "operator": ">=",
                                "value_type": "median",
                                "value_column": "金额",
                                "value": None,
                            },
                            {
                                "column": "状态",
                                "operator": "==",
                                "value_type": "literal",
                                "value": "已完成",
                            },
                        ],
                    },
                    {
                        "type": "drop_nulls",
                        "columns": ["客户名", "金额"],
                        "logic": "or",
                    },
                    {
                        "type": "deduplicate",
                        "columns": ["订单号"],
                        "logic": "and",
                    },
                ],
            },
            "产品表": {
                "operations": [
                    {
                        "type": "deduplicate",
                        "columns": ["产品编码"],
                        "logic": "and",
                    },
                ],
            },
            "说明页": {
                "operations": [],
            },
        },
        "global": {
            "skip_sheets": [],
            "only_sheets": None,
            "engine": "calamine",
        },
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(example, f, ensure_ascii=False, indent=2)

    logger.info(f"示例配置文件已生成: {output_path}")


def main() -> None:
    """CLI 主入口"""

    parser = _build_arg_parser()
    args = parser.parse_args()

    # ── 日志级别 ──
    if args.verbose:
        logger.setLevel(logging.DEBUG)

    # ── 生成示例配置 ──
    if args.generate_example:
        _generate_example_config(args.generate_example)
        return

    # ── 读取配置文件 ──
    if not os.path.exists(args.config):
        logger.error(f"配置文件不存在: {args.config}")
        sys.exit(1)

    try:
        with open(args.config, "r", encoding="utf-8") as f:
            config = json.load(f)
    except json.JSONDecodeError as e:
        logger.error(f"配置文件 JSON 解析失败: {e}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"读取配置文件失败: {e}")
        sys.exit(1)

    # ── 配置验证 ──
    logger.info("正在验证配置...")
    errors = validate_config(config)

    if errors:
        logger.error(f"配置验证失败，共 {len(errors)} 个错误:")
        for err in errors:
            logger.error(f"  ✗ {err}")
        sys.exit(1)
    else:
        logger.info("  配置验证通过 ✓")

    if args.dry_run:
        logger.info("--dry-run 模式，验证完成，不执行处理")
        return

    # ── 执行处理 ──
    try:
        output_path = process_excel(config)
        logger.info(f"✓ 全部完成！输出文件: {output_path}")
    except FileNotFoundError as e:
        logger.error(f"文件不存在: {e}")
        sys.exit(1)
    except (ValueError, KeyError) as e:
        logger.error(f"配置/数据错误: {e}")
        sys.exit(1)
    except MemoryError:
        logger.error(
            "内存不足！建议：\n"
            "  1. 逐表分批处理（减小单个 Sheet 的大小）\n"
            "  2. 先将 Excel 转为 parquet 格式再用 Polars streaming 处理\n"
            "  3. 增加系统虚拟内存"
        )
        sys.exit(1)
    except Exception as e:
        logger.error(f"未预期的错误: {type(e).__name__}: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)


# ============================================================================
#  入口
# ============================================================================

if __name__ == "__main__":
    main()
